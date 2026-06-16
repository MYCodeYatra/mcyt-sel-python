import os
import time
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
def test_excel_export_contents(driver):
    # 1. UI ACTION: Export the data
    driver.get("https://mycodeyatra.com/admin/reports")
    driver.find_element(By.ID, "export-excel-btn").click()
    time.sleep(3)
    excel_path = "/Users/qa/Downloads/Monthly_Report.xlsx"
    assert os.path.exists(excel_path), "Excel Report failed to download!"
    # 2. DOCUMENT ASSERTION: Parse the Excel File
    print(f"\n[Validation] Opening Excel: {excel_path}")
    # Load the workbook
    workbook = load_workbook(excel_path)
    # Assert the correct sheet was generated
    assert "Sales Data" in workbook.sheetnames, "Missing 'Sales Data' sheet!"
    sheet = workbook["Sales Data"]
    # 3. CELL ASSERTIONS
    # Verify headers
    assert sheet['A1'].value == "Transaction ID"
    assert sheet['B1'].value == "Month"
    assert sheet['C1'].value == "Revenue"
    # Verify data in Row 2
    assert sheet['B2'].value == "March 2025"
    # Verify mathematical accuracy by iterating over columns
    total_revenue = 0
    # Assuming data exists from row 2 to 10
    for row in range(2, 11):
        cell_value = sheet[f'C{row}'].value
        if cell_value is not None:
            total_revenue += float(cell_value)
    expected_total = 55000.00
    assert total_revenue == expected_total, f"Excel data does not add up! Expected {expected_total}, Got {total_revenue}"