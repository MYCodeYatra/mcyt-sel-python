import openpyxl
import pytest
def read_excel_data():
    data = []
    workbook = openpyxl.load_workbook('TestData.xlsx')
    sheet = workbook.active
    # Iterate through rows starting from row 2 (skipping header)
    for i in range(2, sheet.max_row + 1):
        username = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        msg = sheet.cell(row=i, column=3).value
        data.append((username, password, msg))
    return data
@pytest.mark.parametrize("username, password, expected_msg", read_excel_data())
def test_login_from_excel(username, password, expected_msg):
    # ... Selenium logic
    pass