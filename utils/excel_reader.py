import openpyxl
def get_data_from_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    # Skip the header row (start from row 2)
    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row, 1).value
        password = sheet.cell(row, 2).value
        expected = sheet.cell(row, 3).value
        # Avoid appending empty rows
        if username is not None:
            data.append((username, password, expected))
    return data